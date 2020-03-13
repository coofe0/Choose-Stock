#! /usr/bin/python3

###目的：爬取符合我投资要求的上市公司的股票名称。学习python###
###路线：网易个股行情网页爬取股票信息，分析财务数据###
###      扩展到整个国内的上实公司，筛选符合我要求的公司###

import re
import pandas as pd
from bs4 import BeautifulSoup
import numpy as np
import bs4
import xlwt
import openpyxl
import os

from getHtml import getHTMLText
from openpyxl import load_workbook

#-------------获取财务数据列表---------------#
def getTableList(html):
	datelist=[]
	navlist=[]
	soup=BeautifulSoup(html,'html.parser')		
	for n in soup.find_all(name='td',string='每股净资产-摊薄/期末股数'):
		t=n.find_next_sibling().string
		navlist.append(t)	
	for n in soup.find_all(name='strong',string='截止日期'):
		t=n.find_parent().find_next_sibling().string
		datelist.append(t)
	print('The length of date is:',len(datelist))
	print('The length of nav is:',len(navlist))
#	print(datelist)
#	print(navlist)

	return datelist,navlist

#---------------------打开并读出文件，转换成list格式------------------------#
def getStockList():
	with open('RoeRolStockList2.txt','r',encoding='utf-8') as fl:       
		t=fl.read()
		lis=re.split('\n',t)	
#		lis=re.split('[\[\],]',t)
		for i in range(len(lis)):
			t=re.search('\d{6}',lis[i])
			if t:
				lis[i]=t.group(0)
	del lis[-1]
	print(lis)
	return lis


#-----------------依次获得每个股票的每股净资产数据并记录-------------------#
def getNAVList():
	l=getStockList()	
	count=0
	for i in range(len(l)):
		print("The stock code is",l[i])
		finance_url='https://vip.stock.finance.sina.com.cn/corp/go.php/vFD_FinanceSummary/stockid/'+l[i]+'.phtml'		#财务数据网址
		finance_html=getHTMLText(finance_url)
		if finance_html !="Wrong":
			datelist,navlist=getTableList(finance_html)			#获得每股净资产数据
			print(navlist)
			for j in range(len(navlist)):
				if navlist[j]:
					try:
						navlist[j]=re.search(r'\d+[.]\d+',navlist[j]).group(0)
					except:
						print('Can not re search:',navlist[j])
				else:
					navlist[j]=0
			navlist.insert(0,'每股净资产')					#插入一个index
			print('navlist is :',len(navlist))		
#			print(navlist)
			wb=load_workbook('all_finance_sheet_done.xlsx')			#打开原财务数据文件
			sheet=wb.get_sheet_by_name(l[i])				#获得sheet
			print('origin is:',sheet.max_column)
			k=0
			for cell in sheet['4']:
				if k==len(navlist):
					break
				cell.value=navlist[k]
				k=k+1
#				print(cell.value)
			print('The sheet have saved:',l[i])
			count+=1
			print('It have saved:',count)
			wb.save('all_finance_sheet_done.xlsx')
			wb.close()			
'''
			lsn.insert(0,'时间')
			npdata=np.array(ls)
			data=pd.DataFrame(npdata,index=lsn)
			print(data.head())

			if os.path.exists('all_finance_sheet.xlsx') == True:
				writer=pd.ExcelWriter('all_finance_sheet.xlsx',mode='a')
				data.to_excel(writer,sheet_name=l[i])
			else:
				writer=pd.ExcelWriter('all_finance_sheet.xlsx')
				data.to_excel(writer,sheet_name=l[i])
			print('The stock finance has write:',l[i])
			writer.save()
	writer.close()


'''



###################---main()---###################
if __name__=="__main__":
	count=0
	getNAVList()
	
